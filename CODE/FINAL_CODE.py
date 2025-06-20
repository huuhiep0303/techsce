#CODE OK 
import sys
import cv2
import face_recognition
import os
import numpy as np
from datetime import datetime
import openpyxl
from PyQt5 import QtWidgets, uic, QtGui, QtCore
from twilio.rest import Client

image_directory = r"D:\KHKT_12_FINALLY\KHKT_12_FINAL\DATA"

class CameraApp(QtWidgets.QMainWindow):
    def __init__(self):
        super(CameraApp, self).__init__()
        uic.loadUi(r'D:\KHKT_12_FINALLY\KHKT_12_FINAL\FILE_UI\Diemdanh_UI.ui', self)  

        self.video_capture = None
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.update_frame)
        self.setFixedSize(1291, 734)

        self.images = []
        self.known_names = []
        self.encoded_known_faces = []
        self.excel_file = None  

        self.load_images()
        self.encode_faces()

        # Kết nối các sự kiện của nút
        self.btnStartStop.clicked.connect(self.toggle_camera)
        self.btnSelectExcel.clicked.connect(self.select_excel_file)
        self.btnShowMissing.clicked.connect(self.show_missing_attendance)
        self.btnShowStudent.clicked.connect(self.show_attendance)
        self.btnClear.clicked.connect(self.functionclear)
        
        # Kết nối QDateEdit với sự kiện chọn ngày
        self.dateEditAttendance.dateChanged.connect(self.check_selected_date)
        
        # Kết nối sự kiện khi nhấn nút hiển thị ngày đã chọn
        self.btnShowSelectedDate.clicked.connect(self.show_selected_date)

        self.textEdit_missingStudents.setReadOnly(True)
        self.NameExcel.setReadOnly(True)

        # Lưu ngày chọn
        self.selected_date = None  

    def load_images(self):
        image_list = os.listdir(image_directory)
        for image_file in image_list:
            current_image = cv2.imread(f"{image_directory}/{image_file}")
            if current_image is not None:
                self.images.append(current_image)
                self.known_names.append(os.path.splitext(image_file)[0])

    def encode_faces(self):
        for img in self.images:
            rgb_image = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encodings = face_recognition.face_encodings(rgb_image)
            if encodings:
                self.encoded_known_faces.append(encodings[0])

    def check_selected_date(self):
        self.selected_date = self.dateEditAttendance.date().toString("dd/MM/yyyy")

    def show_selected_date(self):
        if self.selected_date:
            QtWidgets.QMessageBox.information(self, "Ngày đã chọn", f"Bạn đã chọn ngày: {self.selected_date}")

            if not self.excel_file:
                QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn file Excel.")
                return

            try:
                workbook = openpyxl.load_workbook(self.excel_file)
                sheet = workbook.active
                
                column_letter = None
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value == self.selected_date:
                        column_letter = col
                        break

                if column_letter is None:
                    column_letter = sheet.max_column + 1
                    sheet.cell(row=1, column=column_letter).value = self.selected_date
                else:
                    None

                workbook.save(self.excel_file)
                workbook.close()

            except Exception as e:
                print(f"Lỗi khi ghi vào Excel: {e}")
                QtWidgets.QMessageBox.critical(self, "Lỗi", "Đã xảy ra lỗi trong quá trình ghi vào Excel.")
        else:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn ngày. Vui lòng chọn ngày từ QDateEdit.")

    # def toggle_camera(self):
    #     if not self.excel_file:
    #         QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn file Excel.")
    #         return 
    #     camera_index = self.ID_Edit.toPlainText().strip() 
    
    #     try: 
    #         camera_index = int(camera_index)
    #     except ValueError: 
    #         QtWidgets.QMessageBox.warning(self, "Lỗi", "Giá trị ID không hợp lệ")
    #         return
        
    #     if self.video_capture is None: 
    #         self.video_capture = cv2.VideoCapture(camera_index) 
    #         if not self.video_capture.isOpened():
    #             print("Không thể mở camera")
    #             return
    #         self.btnStartStop.setText("Tắt Camera") 
    #         self.timer.start(20) 
    #     else: 
    #         self.timer.stop()  
    #         self.video_capture.release()
    #         self.Label_Camera.clear()  
    #         self.video_capture = None 
    #         self.btnStartStop.setText("Bật Camera") 
    
    
    def toggle_camera(self):
        if not self.excel_file:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn file Excel.")
            return 
    
        if self.video_capture is None: 
            self.video_capture = cv2.VideoCapture(0) 
            if not self.video_capture.isOpened():
                print("Không thể mở camera")
                return
            self.btnStartStop.setText("Tắt Camera") 
            self.timer.start(20) 
        else: 
            self.timer.stop()  
            self.video_capture.release()
            self.Label_Camera.clear()  
            self.video_capture = None 
            self.btnStartStop.setText("Bật Camera") 

    def update_frame(self):
        ret, frame = self.video_capture.read()
        if not ret:
            print("Không thể đọc khung hình từ webcam")
            return

        small_frame = cv2.resize(frame, (0, 0), None, fx=0.5, fy=0.5)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        current_face_locations = face_recognition.face_locations(rgb_small_frame)
        current_face_encodings = face_recognition.face_encodings(rgb_small_frame, current_face_locations)

        for face_encoding, face_location in zip(current_face_encodings, current_face_locations):
            matches = face_recognition.compare_faces(self.encoded_known_faces, face_encoding)
            face_distances = face_recognition.face_distance(self.encoded_known_faces, face_encoding)
            best_match_index = np.argmin(face_distances)

            if face_distances[best_match_index] < 0.50:
                name = self.known_names[best_match_index].lower()
                self.mark_attendance(name)
            else:
                name = "Unknown"

            top, right, bottom, left = face_location
            top, right, bottom, left = top * 2, right * 2, bottom * 2, left * 2
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = frame.shape
        bytes_per_line = ch * w
        qt_image = QtGui.QImage(frame.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        self.Label_Camera.setPixmap(QtGui.QPixmap.fromImage(qt_image))

    def select_excel_file(self):
        options = QtWidgets.QFileDialog.Options()
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Chọn file Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            self.excel_file = file_name
            base_file_name = os.path.basename(self.excel_file)
            self.NameExcel.setText(base_file_name)
              
    def mark_attendance(self, name):
        if not self.excel_file:
            return
        if not self.selected_date:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn ngày điểm danh.")
            return
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active
            column_letter = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value == self.selected_date:
                    column_letter = col
                    break
            if column_letter is None:
                column_letter = sheet.max_column + 1
                sheet.cell(row=1, column=column_letter).value = self.selected_date
                print(f"Thêm cột mới cho ngày: {self.selected_date}")
            else:
                None

            now = datetime.now()
            datetime_string = now.strftime("%H:%M:%S")
            updated = False
            already_marked = False

            for row in sheet.iter_rows(min_row=2, max_col=column_letter):
                cell_name = row[0].value
                if cell_name and cell_name.strip().lower() == name.strip().lower():
                    if row[column_letter-1].value:
                        already_marked = True
                        break
                    else:
                        sheet.cell(row=row[0].row, column=column_letter).value = f"Đã điểm danh - {datetime_string}"
                        updated = True
                        break

            if already_marked:
                self.textEdit_missingStudents.setText(f"Học sinh {name} đã điểm danh vào ngày {self.selected_date}.")
            elif not updated:
                self.textEdit_missingStudents.setText(f"Học sinh {name} không có trong danh sách lớp này.")
                
            workbook.save(self.excel_file)
            workbook.close()
        except Exception as e:
            print(f"Lỗi trong quá trình ghi vào Excel: {e}")

    def show_attendance(self):
        if not self.excel_file:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn file Excel")
            return 
        
        if not self.selected_date:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn ngày điểm danh.")
            return 
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active 
            attendence_student = []
            
            column_letter = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == self.selected_date:
                    column_letter = col
                    break
            if column_letter is None:
                self.textEdit_missingStudents.setText(f"Chưa có học sinh nào điểm danh vào ngày {self.selected_date}.")
                workbook.close()
                return

            for row in sheet.iter_rows(min_row=2, max_col=column_letter):
                cell_name = row[0].value
                if cell_name and row[column_letter-1].value: 
                    attendence_student.append(cell_name)
            
            if attendence_student:
                missing_students_text = "\n".join(attendence_student)
                self.textEdit_missingStudents.setText(f"Học sinh đã điểm danh vào ngày {self.selected_date}:\n{missing_students_text}")
            else:
                self.textEdit_missingStudents.setText(f"Không có học sinh nào điểm danh vào ngày {self.selected_date}.")
            
            workbook.close()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", "Đã xảy ra lỗi khi đọc file Excel.")

    def show_missing_attendance(self):
        if not self.excel_file:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn file Excel")
            return 
        
        if not self.selected_date:
            QtWidgets.QMessageBox.warning(self, "Lỗi", "Chưa chọn ngày điểm danh.")
            return 

        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active 
            missing_students = []
            
            column_letter = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == self.selected_date:
                    column_letter = col
                    break
            if column_letter is None:
                self.textEdit_missingStudents.setText(f"Chưa có học sinh nào điểm danh vào ngày {self.selected_date}.")
                workbook.close()
                return
            for row in sheet.iter_rows(min_row=2, max_col=column_letter):
                cell_name = row[0].value
                if cell_name:
                    if not row[column_letter-1].value: 
                        missing_students.append(cell_name)

            if missing_students:
                missing_students_text = "\n".join(missing_students)
                self.textEdit_missingStudents.setText(f"Học sinh chưa điểm danh vào ngày {self.selected_date}:\n{missing_students_text}")
            else:
                self.textEdit_missingStudents.setText(f"Tất cả học sinh đã điểm danh vào ngày {self.selected_date}.")
            
            workbook.close()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Lỗi", "Đã xảy ra lỗi khi đọc file Excel.")

    def functionclear(self):
        self.textEdit_missingStudents.clear()
        
    def send_sms_twilio(self, phone_number, message):
        account_sid = 'my_ account_sid'
        auth_token = 'my_auth_token'
        twilio_number = 'my_twilio_number'
    
        client = Client(account_sid, auth_token)
        client.messages.create(
            body = message,
            from_=twilio_number,
            to=phone_number
        )
        print(f"Đã gửi tin nhắn cho {phone_number}.")
    
    def load_parent_phonenumbers(self):
        self.parent_phonenumbers = {}
        wb = openpyxl.load_workbook(r'D:\KHKT_12_FINALLY\KHKT_12_FINAL\EXCEL\parent_phonenumbers.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_name, phone_number = row
            self.parent_phonenumbers[student_name] = phone_number
            
    def get_missing_students(self):
        all_students = self.known_names
        attended_students = self.get_attended_students()
        missing_students= list(set(all_students) - set(attended_students))
        return missing_students
    
    def send_sms(self):
        if not self.excel_file:
            print("Chưa chọn file Excel")
            return
    
        missing_students = self.get_missing_students()
        if missing_students:
            for student in missing_students:
                student_key = student.strip().lower()
                phone = self.parent_phonenumbers.get(student_key)
                if phone:
                    message = f"Học sinh {student} không có mặt trong buổi học ngày {self.selected_date}."
                    try:
                        self.send_sms_twilio(phone, message)
                        print(f"Đã gửi tin nhắn cho {phone}.")
                    except Exception as e:
                        print(f"Lỗi khi gửi tin nhắn: {e}")
            self.textEdit_missingStudents.setPlainText('\n'.join(missing_students))
        else:
            self.textEdit_missingStudents.setPlainText("Tất cả học sinh đã điểm danh.")
        
    def get_attended_students(self):
        attended = []
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if name:
                attended.append(name)
        return attended

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = CameraApp()
    window.show()
    sys.exit(app.exec_())